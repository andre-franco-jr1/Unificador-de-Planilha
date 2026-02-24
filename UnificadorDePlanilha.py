from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from copy import copy
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import sys
import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TARGET_SHEET_NAMES = [
    "Composições com Preço Unitário",
    "Curva ABC de Insumos",
    "Orçamento Sintético",
]


# ============================================================================
# FUNÇÕES UTILITÁRIAS
# ============================================================================

def normalize_description(text):
    """Normaliza descrições para comparação confiável (equivalente ao Trim do VBA + limpeza extra)"""
    if not text or text == 'None':
        return ''
    text = str(text).strip()
    text = re.sub(r'\s+', ' ', text)
    text = text.upper()
    return text


def normalize_sintetico_coluna_a(value):
    """
    Normaliza o código hierárquico da Coluna A do Orçamento Sintético.
    Regras:
    - Remove espaços
    - Troca vírgula por ponto
    - Mantém valor como texto para evitar influência de locale no Excel
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, 'g')
    text_value = str(value).replace(' ', '')
    if text_value == '':
        return ''
    text_value = text_value.replace(',', '.')
    return text_value


def copy_cell_style_and_value(source_cell, target_cell):
    """Copia valor e todo o estilo de uma célula para outra."""
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def clear_cell(cell):
    """Limpa o valor e o estilo de uma célula, voltando ao padrão."""
    cell.value = None
    cell.style = 'Normal'
    cell.fill = PatternFill(fill_type=None)
    cell.border = Border()


def get_cell_color(cell):
    """Retorna o código Hex da cor de fundo de uma célula."""
    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
        color_rgb = str(cell.fill.start_color.rgb)
        if len(color_rgb) == 8:
            return color_rgb[2:].upper()
        elif len(color_rgb) == 6:
            return color_rgb.upper()
    return None


# ============================================================================
# FUNÇÕES DE PROCESSAMENTO DE PLANILHAS
# ============================================================================

def unmerge_all_cells(workbook):
    """ETAPA 2: Remove todas as células mescladas de todas as planilhas do workbook"""
    print('\n==================================================')
    print('ETAPA 2: REMOVENDO MESCLAR E CENTRALIZAR')
    print('==================================================')
    for sheet in workbook.worksheets:
        print(f'  Processando sheet: {sheet.title}')
        merged_ranges = list(sheet.merged_cells.ranges)
        if len(merged_ranges) == 0:
            print('    Nenhuma célula mesclada encontrada')
            continue
        print(f'    Encontradas {len(merged_ranges)} células mescladas')
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))
        print(f'    ✓ {len(merged_ranges)} células desmescladas')
    print('✓ Todas as células foram desmescladas')
    print('==================================================')


def process_curva_abc_sheet(workbook):
    """ETAPA 3: Processa a planilha "Curva ABC de Insumos" """
    print('\n==================================================')
    print('ETAPA 3: PROCESSANDO CURVA ABC DE INSUMOS')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Curva ABC de Insumos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Curva ABC de Insumos\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    print(f'  Dimensões iniciais: {target_sheet.max_row} linhas x {target_sheet.max_column} colunas')
    
    columns_to_remove = [('M', 13), ('L', 12), ('J', 10), ('H', 8), ('C', 3)]
    print('\n  Removendo colunas...')
    for col_letter, col_index in columns_to_remove:
        print(f'    Removendo coluna {col_letter}...')
        target_sheet.delete_cols(col_index, 1)
        print(f'      ✓ Coluna {col_letter} removida')
    
    print('\n  Removendo linha 5...')
    target_sheet.delete_rows(5, 1)
    print('    ✓ Linha 5 removida')
    
    print('\n  Alterando célula I4...')
    target_sheet['I4'].value = 'Porcentagem (%)'
    print('    ✓ I4 alterado')
    
    print('\n  Ativando \'Quebra de Texto Automática\' na linha 2...')
    wrap_alignment = Alignment(wrap_text=True)
    for col_idx in range(1, target_sheet.max_column + 1):
        cell = target_sheet.cell(2, col_idx)
        if cell.value:
            current_alignment = cell.alignment
            if current_alignment:
                cell.alignment = Alignment(
                    wrap_text=True,
                    horizontal=current_alignment.horizontal,
                    vertical=current_alignment.vertical,
                    text_rotation=current_alignment.text_rotation,
                    indent=current_alignment.indent
                )
            else:
                cell.alignment = wrap_alignment
    print('    ✓ Quebra de texto ativada na linha 2 da Curva ABC')
    
    print('\n  Copiando coluna G para coluna J...')
    max_row = target_sheet.max_row
    for row in range(1, max_row + 1):
        source_cell = target_sheet.cell(row, 7)
        target_cell = target_sheet.cell(row, 10)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    print('    ✓ Coluna copiada')
    
    print('\n  Alterando célula J4...')
    target_sheet['J4'].value = 'Valor unitário (BASE)'
    print('    ✓ J4 alterado')
    
    print('\n  Alterando célula L4...')
    target_sheet['L4'].value = 'Observações'
    print('    ✓ L4 alterado')
    
    print('\n  Convertendo texto para número...')
    columns_to_convert = [6, 7, 8, 9, 10, 11, 12]
    for row in range(1, max_row + 1):
        for col in columns_to_convert:
            cell = target_sheet.cell(row, col)
            if cell.value is not None:
                try:
                    if isinstance(cell.value, str):
                        clean_value = cell.value.strip().replace(',', '.')
                        numeric_value = float(clean_value)
                        cell.value = numeric_value
                except (ValueError, AttributeError):
                    pass
    print('    ✓ Conversão concluída')
    
    print('\n  Movendo D1 e D2 para E1 e E2...')
    d1_cell = target_sheet['D1']
    e1_cell = target_sheet['E1']
    e1_cell.value = d1_cell.value
    if d1_cell.has_style:
        e1_cell.font = copy(d1_cell.font)
        e1_cell.border = copy(d1_cell.border)
        e1_cell.fill = copy(d1_cell.fill)
        e1_cell.number_format = copy(d1_cell.number_format)
        e1_cell.protection = copy(d1_cell.protection)
        e1_cell.alignment = copy(d1_cell.alignment)
    d1_cell.value = None
    
    d2_cell = target_sheet['D2']
    e2_cell = target_sheet['E2']
    e2_cell.value = d2_cell.value
    if d2_cell.has_style:
        e2_cell.font = copy(d2_cell.font)
        e2_cell.border = copy(d2_cell.border)
        e2_cell.fill = copy(d2_cell.fill)
        e2_cell.number_format = copy(d2_cell.number_format)
        e2_cell.protection = copy(d2_cell.protection)
        e2_cell.alignment = copy(d2_cell.alignment)
    d2_cell.value = None
    print('    ✓ Movimentação concluída')
    
    target_sheet['D1'].value = 'Bancos'
    target_sheet['D1'].font = Font(bold=False)
    print('    ✓ D1 = \'Bancos\'')
    
    target_sheet['I1'].value = 'UNIBASE'
    target_sheet['I1'].font = Font(bold=False)
    target_sheet['J1'].value = 'CERTAME'
    target_sheet['J1'].font = Font(bold=False)
    target_sheet['K1'].value = 'DESCONTO'
    target_sheet['K1'].font = Font(bold=False)
    target_sheet['L1'].value = 'ACERVO'
    target_sheet['L1'].font = Font(bold=False)
    target_sheet['F1'].value = 'B.D.I.'
    target_sheet['F1'].font = Font(bold=False)
    
    target_sheet['J2'].value = 1
    target_sheet['K2'].value = '=1-I2/J2'
    target_sheet['L2'].value = 'RECIFE-20XX'
    
    target_sheet.merge_cells('A3:L3')
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=False)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_a3 = target_sheet['A3']
    cell_a3.fill = blue_fill
    cell_a3.font = white_font
    cell_a3.alignment = center_alignment
    
    thin_border_side = Side(style='thin')
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    
    last_row = target_sheet.max_row
    for row_idx in range(1, last_row + 1):
        for col_idx in range(1, 13):
            cell = target_sheet.cell(row_idx, col_idx)
            cell.border = thin_border
    
    target_sheet['I2'].number_format = 'R$ #,##0.00'
    for row_idx in range(1, last_row + 1):
        target_sheet.cell(row_idx, 7).number_format = 'R$ #,##0.00'
        target_sheet.cell(row_idx, 8).number_format = 'R$ #,##0.00'
        target_sheet.cell(row_idx, 10).number_format = 'R$ #,##0.00'
    
    target_sheet['F2'].number_format = '0.00%'
    for row_idx in range(1, last_row + 1):
        target_sheet.cell(row_idx, 11).number_format = '0.00%'
    
    for row_idx in range(1, last_row + 1):
        if row_idx != 2:
            target_sheet.cell(row_idx, 9).number_format = '0.00%'
    
    # Aplicar formatação de milhar nas colunas F, I, K, L
    print('\n  Aplicando formatação de milhar nas colunas numéricas...')
    for row_idx in range(1, last_row + 1):
        if row_idx != 2:  # Não aplicar na linha 2 ainda
            target_sheet.cell(row_idx, 6).number_format = '#,##0.00'  # Coluna F
        target_sheet.cell(row_idx, 9).number_format = '#,##0.00' if row_idx == 2 else '0.00%'  # Coluna I
        target_sheet.cell(row_idx, 12).number_format = '#,##0.00'  # Coluna L
    print('    ✓ Formatação de milhar aplicada')
    
    arial_bold_font = Font(name='Arial', size=11, bold=True)
    center_middle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row_idx in range(1, 3):
        for col_idx in range(4, 13):
            cell = target_sheet.cell(row_idx, col_idx)
            cell.font = arial_bold_font
            cell.alignment = center_middle_alignment
    
    print('\n  Aplicando formatação na linha 4 (negrito e centralizado)...')
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(1, target_sheet.max_column + 1):
        cell = target_sheet.cell(4, col_idx)
        if cell.value:
            cell.font = bold_font
            cell.alignment = center_alignment
    print('    ✓ Linha 4 formatada: negrito e centralizado')
    
    no_fill = PatternFill(fill_type=None)
    for row_idx in range(5, last_row + 1):
        cell_l = target_sheet.cell(row_idx, 12)
        cell_l.value = None
        cell_l.fill = no_fill
    
    # Aplicar quebra de texto específica em G2 e E2 DEPOIS de todas as outras formatações
    print('\n  Aplicando quebra de texto em G2 (encargos) e E2 (bancos)...')
    g2_cell = target_sheet['G2']
    g2_cell.alignment = Alignment(
        wrap_text=True,
        horizontal='center',
        vertical='center'
    )
    e2_cell = target_sheet['E2']
    e2_cell.alignment = Alignment(
        wrap_text=True,
        horizontal='center',
        vertical='center'
    )
    # Aplicar separador de milhar na linha 2 da coluna F
    target_sheet.cell(2, 6).number_format = '#,##0.00'
    print('    ✓ G2 e E2 com quebra de texto aplicada')
    
    print(f'\n  Dimensões finais: {target_sheet.max_row} linhas x {target_sheet.max_column} colunas')
    print('✓ Processamento concluído')
    print('==================================================')


def apply_formulas_to_curva_abc(workbook):
    """ETAPA 4: Aplica fórmulas nas colunas G e H"""
    print('\n==================================================')
    print('ETAPA 4: APLICANDO FÓRMULAS NAS COLUNAS G E H')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Curva ABC de Insumos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    target_colors = ['D6D6D6', 'EFEFEF', 'F7F3DF']
    formulas_g = 0
    formulas_h = 0
    
    for row_idx in range(1, target_sheet.max_row + 1):
        cell_g = target_sheet.cell(row_idx, 7)
        bg_color_g = get_cell_color(cell_g)
        if bg_color_g in target_colors:
            formula_g = f'=J{row_idx}*(1-K{row_idx})'
            cell_g.value = formula_g
            formulas_g += 1
        
        cell_h = target_sheet.cell(row_idx, 8)
        bg_color_h = get_cell_color(cell_h)
        if bg_color_h in target_colors:
            formula_h = f'=F{row_idx}*G{row_idx}'
            cell_h.value = formula_h
            formulas_h += 1
    
    print(f'  ✓ Fórmulas G: {formulas_g}')
    print(f'  ✓ Fórmulas H: {formulas_h}')
    print('✓ Concluído')
    print('==================================================')


def apply_sum_and_formulas_to_curva_abc(workbook):
    """ETAPA 5: Aplica soma na coluna H e fórmulas nas colunas I e K"""
    print('\n==================================================')
    print('ETAPA 5: APLICANDO SOMA EM H E FÓRMULAS EM I E K')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Curva ABC de Insumos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    target_colors = ['D6D6D6', 'EFEFEF', 'F7F3DF']
    colored_rows = []
    last_filled_row = 0
    
    for row_idx in range(1, target_sheet.max_row + 1):
        cell_h = target_sheet.cell(row_idx, 8)
        if cell_h.value is not None:
            last_filled_row = row_idx
        bg_color_h = get_cell_color(cell_h)
        if bg_color_h in target_colors:
            colored_rows.append(row_idx)
    
    sum_row = last_filled_row + 1
    
    if colored_rows:
        ranges = []
        start = colored_rows[0]
        end = colored_rows[0]
        for i in range(1, len(colored_rows)):
            if colored_rows[i] == end + 1:
                end = colored_rows[i]
            else:
                if start == end:
                    ranges.append(f'H{start}')
                else:
                    ranges.append(f'H{start}:H{end}')
                start = colored_rows[i]
                end = colored_rows[i]
        
        if start == end:
            ranges.append(f'H{start}')
        else:
            ranges.append(f'H{start}:H{end}')
        
        sum_formula = f"=SUM({','.join(ranges)})"
        target_sheet.cell(sum_row, 8).value = sum_formula
    
    formulas_i = 0
    formulas_k = 0
    for row_idx in colored_rows:
        cell_i = target_sheet.cell(row_idx, 9)
        bg_color_i = get_cell_color(cell_i)
        if bg_color_i in target_colors:
            formula_i = f'=H{row_idx}/$H${sum_row}'
            cell_i.value = formula_i
            formulas_i += 1
        
        cell_k = target_sheet.cell(row_idx, 11)
        bg_color_k = get_cell_color(cell_k)
        if bg_color_k in target_colors:
            cell_k.value = '=0'
            formulas_k += 1
    
    print(f'  ✓ Fórmulas I: {formulas_i}')
    print(f'  ✓ Fórmulas K: {formulas_k}')
    print('✓ Concluído')
    print('==================================================')


def apply_sum_column_i_and_delete_below(workbook):
    """ETAPA 6: Aplica soma na coluna I e deleta linhas abaixo"""
    print('\n==================================================')
    print('ETAPA 6: APLICANDO SOMA EM I E DELETANDO LINHAS ABAIXO')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Curva ABC de Insumos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    target_colors = ['D6D6D6', 'EFEFEF', 'F7F3DF']
    colored_rows = []
    last_colored_row = 0
    
    for row_idx in range(1, target_sheet.max_row + 1):
        cell_i = target_sheet.cell(row_idx, 9)
        bg_color_i = get_cell_color(cell_i)
        if bg_color_i in target_colors:
            colored_rows.append(row_idx)
            last_colored_row = row_idx
    
    sum_row = last_colored_row + 1
    
    if colored_rows:
        ranges = []
        start = colored_rows[0]
        end = colored_rows[0]
        for i in range(1, len(colored_rows)):
            if colored_rows[i] == end + 1:
                end = colored_rows[i]
            else:
                if start == end:
                    ranges.append(f'I{start}')
                else:
                    ranges.append(f'I{start}:I{end}')
                start = colored_rows[i]
                end = colored_rows[i]
        
        if start == end:
            ranges.append(f'I{start}')
        else:
            ranges.append(f'I{start}:I{end}')
        
        sum_formula = f"=SUM({','.join(ranges)})"
        target_sheet.cell(sum_row, 9).value = sum_formula
    
    total_rows = target_sheet.max_row
    rows_to_delete = total_rows - sum_row
    if rows_to_delete > 0:
        target_sheet.delete_rows(sum_row + 1, rows_to_delete)
    
    # Remover bordas e centralizar células após a linha de soma
    print('\n  Removendo bordas e centralizando células após composições...')
    no_border = Border()
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row_idx in range(sum_row, target_sheet.max_row + 1):
        for col_idx in range(1, 13):
            cell = target_sheet.cell(row_idx, col_idx)
            cell.border = no_border
            cell.alignment = center_alignment
    print('    ✓ Bordas removidas e conteúdo centralizado após composições')
    
    print('✓ Concluído')
    print('==================================================')


def process_cpus_sheet(workbook):
    """ETAPA 7: Processa a planilha CPUs"""
    print('\n==================================================')
    print('ETAPA 7: PROCESSANDO PLANILHA CPUs')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'CPU' in sheet.title or 'Compos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None and len(workbook.worksheets) > 0:
        target_sheet = workbook.worksheets[0]
    if target_sheet is None:
        print('  ⚠️  Sheet não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    target_sheet.delete_rows(1, 2)
    target_sheet.delete_cols(6, 1)
    
    rows_cleared = []
    for row_idx in range(1, target_sheet.max_row + 1):
        cell_e = target_sheet.cell(row_idx, 5)
        if not cell_e.value:
            continue
        
        cell_value = str(cell_e.value).strip()
        if 'MO sem LS =>' in cell_value or 'Valor do BDI =>' in cell_value:
            cell_e.value = None
            rows_cleared.append(row_idx)
    
    hidden_rows = 0
    for row_idx in range(1, target_sheet.max_row + 1):
        if row_idx in target_sheet.row_dimensions:
            if target_sheet.row_dimensions[row_idx].hidden or target_sheet.row_dimensions[row_idx].height == 0:
                target_sheet.row_dimensions[row_idx].hidden = False
                target_sheet.row_dimensions[row_idx].height = None
                hidden_rows += 1
    
    print('✓ Concluído')
    print('==================================================')


def clear_specific_rows_cpus(workbook):
    """ETAPA LIMPEZA CPUs: Limpa linhas inteiras que contêm frases específicas"""
    print('\n==================================================')
    print('ETAPA LIMPEZA: REMOVENDO LINHAS ESPECÍFICAS DA CPUs')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'CPU' in sheet.title or 'Compos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None and len(workbook.worksheets) > 0:
        target_sheet = workbook.worksheets[0]
    if target_sheet is None:
        print('  ⚠️  Sheet \'CPUs\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    phrases_to_clear = ['Valor com BDI =>', 'MO com LS =>']
    rows_to_clear = []
    
    print(f'\n  Procurando linhas com frases: {phrases_to_clear}')
    for row_idx in range(1, target_sheet.max_row + 1):
        for col_idx in range(1, target_sheet.max_column + 1):
            cell = target_sheet.cell(row_idx, col_idx)
            if cell.value:
                cell_value = str(cell.value).strip()
                for phrase in phrases_to_clear:
                    if phrase in cell_value:
                        rows_to_clear.append(row_idx)
                        print(f'    ⚠️  Linha {row_idx} contém \'{phrase}\' → será limpa')
                        break
                if row_idx in rows_to_clear:
                    break
    
    if rows_to_clear:
        print(f'\n  Limpando {len(rows_to_clear)} linha(s)...')
        for row_idx in rows_to_clear:
            for col_idx in range(1, target_sheet.max_column + 1):
                cell = target_sheet.cell(row_idx, col_idx)
                clear_cell(cell)
        print(f'  ✓ {len(rows_to_clear)} linha(s) limpa(s) completamente')
        if len(rows_to_clear) <= 10:
            print(f"  ✓ Linhas limpas: {', '.join(map(str, rows_to_clear))}")
        else:
            print(f"  ✓ Primeiras linhas limpas: {', '.join(map(str, rows_to_clear[:10]))}")
            print(f'  ✓ ... e mais {len(rows_to_clear) - 10} linhas')
    else:
        print('  ℹ️  Nenhuma linha encontrada com as frases especificadas')
    
    print('✓ Limpeza concluída')
    print('==================================================')


def apply_formulas_cpus_sheet(workbook):
    """ETAPA 8: Aplica fórmulas na planilha CPUs"""
    print('\n==================================================')
    print('ETAPA 8: APLICANDO FÓRMULAS NA PLANILHA CPUs')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'CPU' in sheet.title or 'Compos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None and len(workbook.worksheets) > 0:
        target_sheet = workbook.worksheets[0]
    if target_sheet is None:
        print('  ⚠️  Sheet não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    color_green = 'DFF0D8'
    
    unhidden_count = 0
    for row_idx in range(1, target_sheet.max_row + 1):
        has_green = False
        for col_idx in range(1, target_sheet.max_column + 1):
            cell = target_sheet.cell(row_idx, col_idx)
            bg_color = get_cell_color(cell)
            if bg_color == color_green:
                has_green = True
                break
        if has_green:
            target_sheet.row_dimensions[row_idx].hidden = False
            target_sheet.row_dimensions[row_idx].height = 15
            unhidden_count += 1
    
    color_gray1 = 'D6D6D6'
    color_gray2 = 'EFEFEF'
    target_colors_i = {color_green, color_gray1, color_gray2}
    
    last_row = target_sheet.max_row
    formulas_truncar = 0
    for row_idx in range(1, last_row + 1):
        cell_i = target_sheet.cell(row_idx, 9)
        bg_color = get_cell_color(cell_i)
        if bg_color in target_colors_i:
            formula = f'=TRUNC((G{row_idx}*H{row_idx}),2)'
            cell_i.value = formula
            formulas_truncar += 1
    
    formulas_soma = 0
    for row_idx in range(1, last_row + 1):
        cell_h = target_sheet.cell(row_idx, 8)
        bg_color_h = get_cell_color(cell_h)
        if bg_color_h == color_green:
            cell_c = target_sheet.cell(row_idx, 3)
            cell_c_value = str(cell_c.value) if cell_c.value else ''
            
            start_row_sum = 0
            end_row_sum = 0
            
            if 'ORSE' in cell_c_value.upper():
                for j in range(row_idx + 1, last_row + 1):
                    cell_a = target_sheet.cell(j, 1)
                    cell_a_value = str(cell_a.value).strip() if cell_a.value else ''
                    if 'Detalhamento de Cálculo ORSE' in cell_a_value:
                        start_row_sum = j + 2
                        break
                
                if start_row_sum > 0:
                    for k in range(start_row_sum, last_row + 2):
                        if k > last_row:
                            end_row_sum = last_row
                            break
                        cell_a_k = target_sheet.cell(k, 1)
                        cell_a_k_value = str(cell_a_k.value).strip() if cell_a_k.value else ''
                        if cell_a_k_value != 'Insumo' and cell_a_k_value != 'Item':
                            end_row_sum = k - 1
                            break
            else:
                start_row_sum = row_idx + 1
                for j in range(start_row_sum, last_row + 2):
                    if j > last_row:
                        end_row_sum = last_row
                        break
                    cell_a_j = target_sheet.cell(j, 1)
                    cell_a_j_value = str(cell_a_j.value).strip() if cell_a_j.value else ''
                    if (cell_a_j_value != 'Item' and cell_a_j_value != 'Insumo' and 
                        cell_a_j_value != 'Composição Auxiliar'):
                        end_row_sum = j - 1
                        break
            
            if start_row_sum > 0 and end_row_sum >= start_row_sum:
                formula_soma = f'=SUM(I{start_row_sum}:I{end_row_sum})'
                cell_h.value = formula_soma
                formulas_soma += 1
            else:
                cell_h.value = 0
    
    print(f'  ✓ Fórmulas TRUNCAR: {formulas_truncar}')
    print(f'  ✓ Fórmulas SOMA: {formulas_soma}')
    print('✓ Concluído')
    print('==================================================')


def apply_third_step_cpus_sheet(workbook):
    """ETAPA 9: Aplica fórmulas de referência para Composições Auxiliares (terceiropasso)"""
    print('\n==================================================')
    print('ETAPA 9: APLICANDO REFERÊNCIAS PARA COMPOSIÇÕES AUXILIARES')
    print('==================================================')
    try:
        target_sheet = None
        for sheet in workbook.worksheets:
            if 'CPU' in sheet.title or 'Compos' in sheet.title:
                target_sheet = sheet
                break
        if target_sheet is None and len(workbook.worksheets) > 0:
            target_sheet = workbook.worksheets[0]
        if target_sheet is None:
            print('  ⚠️  Sheet não encontrada')
            print('==================================================')
            return
        
        print(f'  ✓ Sheet encontrada: {target_sheet.title}')
        
        color_green = 'DFF0D8'
        color_gray = 'D6D6D6'
        last_row = target_sheet.max_row
        
        print(f'\n  ETAPA 1: Mapeando composições principais (cor #{color_green})...')
        green_map = {}
        for row_idx in range(1, last_row + 1):
            try:
                cell_h = target_sheet.cell(row_idx, 8)
                if get_cell_color(cell_h) == color_green:
                    cell_b = target_sheet.cell(row_idx, 2)
                    description_raw = str(cell_b.value) if cell_b.value else ''
                    description = normalize_description(description_raw)
                    if description:
                        green_map[description] = row_idx
            except Exception as e:
                print(f'    ⚠️  Erro ao processar linha {row_idx} (mapeamento): {e}')
        
        print(f'  ✓ Total de composições principais mapeadas: {len(green_map)}')
        if len(green_map) == 0:
            print('  ⚠️  Nenhuma composição verde encontrada')
            print('==================================================')
            return
        
        print(f'\n  ETAPA 2: Aplicando fórmulas em composições auxiliares (cor #{color_gray})...')
        formulas_applied = 0
        duplicates_avoided = 0
        
        for row_idx in range(1, last_row + 1):
            try:
                cell_h = target_sheet.cell(row_idx, 8)
                if get_cell_color(cell_h) != color_gray:
                    continue
                
                cell_b = target_sheet.cell(row_idx, 2)
                description_raw = str(cell_b.value) if cell_b.value else ''
                description = normalize_description(description_raw)
                
                if not description or description not in green_map:
                    continue
                
                target_row = green_map[description]
                
                if target_row == row_idx:
                    duplicates_avoided += 1
                    continue
                
                cell_h.value = f'=I{target_row}'
                formulas_applied += 1
            except Exception as e:
                print(f'    ⚠️  Erro ao processar linha {row_idx} (aplicação): {e}')
        
        print(f'  ✓ Total de fórmulas aplicadas: {formulas_applied}')
        print(f'  ✓ Auto-referências evitadas: {duplicates_avoided}')
        print('✓ Terceiro passo (ETAPA 9) concluído com sucesso!')
        print('==================================================')
    except Exception as e:
        print('\n  ✗✗✗ ERRO NA ETAPA 9 ✗✗✗')
        print(f'  Erro: {str(e)}')
        import traceback
        traceback.print_exc()
        print('  ⚠️  Continuando com o processamento...')
        print('==================================================')


def _find_problematic_abc_codes(abc_sheet):
    """Varre a Curva ABC e retorna um set() de códigos problemáticos."""
    print('\n  [Helper] Mapeando códigos problemáticos na Curva ABC...')
    code_map = {}
    problematic_codes = set()
    last_row_abc = abc_sheet.max_row
    
    for row_idx in range(2, last_row_abc + 1):
        try:
            cell_a = abc_sheet.cell(row_idx, 1)
            code = str(cell_a.value).strip() if cell_a.value else ''
            
            cell_b = abc_sheet.cell(row_idx, 2)
            bank = str(cell_b.value).strip() if cell_b.value else ''
            
            if code and code != 'None':
                if code in code_map:
                    if code_map[code] != bank:
                        problematic_codes.add(code)
                else:
                    code_map[code] = bank
        except Exception as e:
            print(f'    ⚠️  Erro ao processar linha {row_idx} da ABC: {e}')
    
    print(f'  [Helper] ✓ {len(problematic_codes)} códigos problemáticos encontrados.')
    return problematic_codes


def apply_fourth_step_cpus_sheet(workbook, problematic_codes):
    """ETAPA 10: Busca de insumos na planilha Curva ABC (quartopasso)"""
    print('\n==================================================')
    print('ETAPA 10: BUSCA DE INSUMOS NA CURVA ABC (MODO COMPATÍVEL)')
    print('==================================================')
    try:
        cpus_sheet = None
        abc_sheet = None
        for sheet in workbook.worksheets:
            if 'CPU' in sheet.title or 'Compos' in sheet.title:
                cpus_sheet = sheet
            elif 'Curva ABC de Insumos' in sheet.title:
                abc_sheet = sheet
        
        if cpus_sheet is None:
            print('  ⚠️  Planilha \'CPUs\' não encontrada')
            print('==================================================')
            return
        if abc_sheet is None:
            print('  ⚠️  Planilha \'Curva ABC de Insumos\' não encontrada')
            print('==================================================')
            return
        
        print(f'  ✓ CPUs: {cpus_sheet.title}')
        print(f'  ✓ Curva ABC: {abc_sheet.title}')
        
        color_gray = 'EFEFEF'
        print(f'  ✓ Usando {len(problematic_codes)} códigos problemáticos pré-calculados.')
        
        print('\n  ETAPA 2: Inserindo fórmulas (VLOOKUP e INDEX/MATCH) na planilha CPUs...')
        last_row_cpus = cpus_sheet.max_row
        formulas_normal = 0
        formulas_problematic = 0
        abc_sheet_name = f'\'{abc_sheet.title}\''
        
        for row_idx in range(1, last_row_cpus + 1):
            try:
                cell_h = cpus_sheet.cell(row_idx, 8)
                bg_color = get_cell_color(cell_h)
                if bg_color == color_gray:
                    cell_b = cpus_sheet.cell(row_idx, 2)
                    code = str(cell_b.value).strip() if cell_b.value else ''
                    
                    if code in problematic_codes:
                        cell_d_ref = f'D{row_idx}'
                        formula_en = f'=IFERROR(INDEX({abc_sheet_name}!$G:$G,MATCH({cell_d_ref},{abc_sheet_name}!$C:$C,0)),"Descrição não encontrada")'
                        cell_h.value = formula_en
                        formulas_problematic += 1
                    else:
                        cell_b_ref = f'B{row_idx}'
                        formula_en = f'=IFERROR(VLOOKUP({cell_b_ref},{abc_sheet_name}!$A:$G,7,FALSE),"Código não encontrado")'
                        cell_h.value = formula_en
                        formulas_normal += 1
            except Exception as e:
                print(f'    ⚠️  Erro ao processar linha {row_idx} da CPUs: {e}')
        
        print(f'\n  ✓ Total de fórmulas VLOOKUP (por código): {formulas_normal}')
        print(f'  ✓ Total de fórmulas INDEX/MATCH (por descrição): {formulas_problematic}')
        print('✓ Quarto passo (ETAPA 10) concluído com sucesso!')
        print('==================================================')
    except Exception as e:
        print('\n  ✗✗✗ ERRO NA ETAPA 10 ✗✗✗')
        print(f'  Erro: {str(e)}')
        import traceback
        traceback.print_exc()
        print('  ⚠️  Continuando com o processamento...')
        print('==================================================')


def apply_fifth_step_cpus_sheet(workbook):
    """ETAPA 11: Limpa conteúdo de linhas verdes com coluna D vazia (quintopasso)"""
    print('\n==================================================')
    print('ETAPA 11: LIMPANDO LINHAS VERDES COM COLUNA D VAZIA')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'CPU' in sheet.title or 'Compos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None and len(workbook.worksheets) > 0:
        target_sheet = workbook.worksheets[0]
    if target_sheet is None:
        print('  ⚠️  Sheet \'CPUs\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    color_green = 'DFF0D8'
    last_row = target_sheet.max_row
    rows_cleared_count = 0
    
    print(f'  ✓ Verificando {last_row} linhas (cor alvo: {color_green})...')
    for row_idx in range(last_row, 0, -1):
        try:
            cell_a = target_sheet.cell(row=row_idx, column=1)
            cell_d = target_sheet.cell(row=row_idx, column=4)
            bg_color = get_cell_color(cell_a)
            
            if bg_color == color_green:
                cell_d_value = cell_d.value
                is_empty = cell_d_value is None or (isinstance(cell_d_value, str) and cell_d_value.strip() == '')
                
                if is_empty:
                    for col_idx in range(1, target_sheet.max_column + 1):
                        target_sheet.cell(row=row_idx, column=col_idx).value = None
                    rows_cleared_count += 1
        except Exception as e:
            print(f'    ⚠️  Erro ao processar linha {row_idx}: {e}')
    
    print(f'\n  ✓ Total de linhas com conteúdo limpo: {rows_cleared_count}')
    print('✓ Concluído')
    print('==================================================')


def apply_sixth_step_cpus_sheet(workbook):
    """ETAPA 12: Aplica formatação de cabeçalho e limpa linhas extras (com fonte Arial)"""
    print('\n==================================================')
    print('ETAPA 12: FORMATANDO CABEÇALHOS E LIMPANDO FINAL DA CPU')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'CPU' in sheet.title or 'Compos' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None and len(workbook.worksheets) > 0:
        target_sheet = workbook.worksheets[0]
    if target_sheet is None:
        print('  ⚠️  Sheet \'CPUs\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    white_font_14_arial = Font(name='Arial', color='FFFFFF', size=14)
    white_font_14 = Font(color='FFFFFF', size=14)
    
    try:
        target_sheet.merge_cells('A1:I1')
        cell_a1 = target_sheet['A1']
        cell_a1.fill = blue_fill
        cell_a1.font = white_font_14_arial
        cell_a1.alignment = center_alignment
        print('  ✓ Linha 1 formatada (A1:I1) com fonte Arial')
    except Exception as e:
        print(f'    ⚠️  Erro ao formatar linha 1: {e}')
    
    try:
        target_sheet.merge_cells('A2:I2')
        cell_a2 = target_sheet['A2']
        cell_a2.fill = blue_fill
        cell_a2.font = white_font_14_arial
        cell_a2.alignment = center_alignment
        print('  ✓ Linha 2 formatada (A2:I2) com fonte Arial')
    except Exception as e:
        print(f'    ⚠️  Erro ao formatar linha 2: {e}')
    
    found_composicao = False
    try:
        for row in target_sheet.iter_rows(min_col=1, max_col=1, min_row=3):
            cell = row[0]
            if cell.value and 'Composições Auxiliares' in str(cell.value):
                row_idx = cell.row
                merge_range = f'A{row_idx}:I{row_idx}'
                target_sheet.merge_cells(merge_range)
                cell.fill = blue_fill
                cell.font = white_font_14
                cell.alignment = center_alignment
                print(f'  ✓ Linha {row_idx} (\'Composições Auxiliares\') formatada')
                found_composicao = True
                break
        if not found_composicao:
            print('  ℹ️  Linha \'Composições Auxiliares\' não encontrada para formatar')
    except Exception as e:
        print(f'    ⚠️  Erro ao formatar linha \'Composições Auxiliares\': {e}')
    
    target_colors = {'EFEFEF', 'DFF0D8', 'D6D6D6'}
    last_colored_row = 0
    max_row = target_sheet.max_row
    
    print(f'\n  ✓ Procurando última linha com cor {target_colors} em {max_row} linhas...')
    try:
        for row_idx in range(1, max_row + 1):
            cell = target_sheet.cell(row=row_idx, column=1)
            bg_color = get_cell_color(cell)
            if bg_color in target_colors:
                last_colored_row = row_idx
        
        if last_colored_row > 0 and last_colored_row < max_row:
            rows_to_delete = max_row - last_colored_row
            print(f'  ✓ Última linha com cor encontrada: {last_colored_row}')
            print(f'  ✓ Deletando {rows_to_delete} linhas (da {last_colored_row + 1} até {max_row})...')
            target_sheet.delete_rows(last_colored_row + 1, rows_to_delete)
            print('  ✓ Linhas extras deletadas.')
        elif last_colored_row == max_row:
            print('  ✓ Última linha com cor já é a última linha da planilha. Nada a deletar.')
        else:
            print('  ℹ️  Nenhuma linha com as cores alvo foi encontrada. Nenhuma linha deletada.')
    except Exception as e:
        print(f'    ⚠️  Erro ao limpar linhas extras da CPU: {e}')
    
    print('✓ Concluído')
    print('==================================================')


def get_item_level_py(item_value):
    """Função auxiliar para determinar o nível hierárquico de um item"""
    if not item_value:
        return -1
    clean_string = str(item_value).strip()
    if clean_string == '':
        return -1
    level = clean_string.count('.') + 1
    return level


def process_sintetico_sheet(workbook):
    """ETAPA 13: Processa a planilha "Orçamento Sintético" (Parte 1)"""
    print('\n==================================================')
    print('ETAPA 13: PROCESSANDO ORÇAMENTO SINTÉTICO (PARTE 1)')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    try:
        h1_cell = target_sheet['H1']
        g1_cell = target_sheet['G1']
        copy_cell_style_and_value(h1_cell, g1_cell)
        clear_cell(h1_cell)
        
        h2_cell = target_sheet['H2']
        g2_cell = target_sheet['G2']
        copy_cell_style_and_value(h2_cell, g2_cell)
        clear_cell(h2_cell)
        print('  ✓ Células H1, H2 movidas para G1, G2.')
        
        print('\n  Ativando \'Quebra de Texto Automática\' na linha 2...')
        wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for col_idx in range(1, target_sheet.max_column + 1):
            cell = target_sheet.cell(2, col_idx)
            if cell.value:
                current_alignment = cell.alignment
                if current_alignment:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=current_alignment.horizontal or 'center',
                        vertical=current_alignment.vertical or 'center',
                        text_rotation=current_alignment.text_rotation,
                        indent=current_alignment.indent
                    )
                else:
                    cell.alignment = wrap_alignment
        print('    ✓ Quebra de texto ativada na linha 2')
    except Exception as e:
        print(f'    ⚠️  Erro ao mover células H1/H2: {e}')
    
    try:
        target_sheet.delete_rows(5, 1)
        print('  ✓ Linha 5 deletada.')
        
        target_sheet.delete_cols(14, 1)
        target_sheet.delete_cols(13, 1)
        target_sheet.delete_cols(10, 1)
        target_sheet.delete_cols(9, 1)
        target_sheet.delete_cols(8, 1)
        print('  ✓ Colunas H, I, J, M, N deletadas.')
    except Exception as e:
        print(f'    ⚠️  Erro ao deletar linhas/colunas: {e}')
    
    try:
        target_sheet['E4'].value = 'Bancos'
        target_sheet['F4'].value = 'B.D.I.'
        target_sheet['G4'].value = 'Valor Unit'
        target_sheet['H4'].value = 'Valor Unit com BDI'
        target_sheet['I4'].value = 'Total sem BDI'
        target_sheet['J4'].value = 'Total com BDI'
        target_sheet['K4'].value = 'Peso (%)'
        target_sheet['K1'].value = 'ACERVO'
        target_sheet['K2'].value = 'RECIFE-20XX'
        print('  ✓ Cabeçalhos atualizados')
        
    except Exception as e:
        print(f'    ⚠️  Erro ao atualizar cabeçalhos do Sintético: {e}')
    
    # Aplicar quebra de texto em H2 e E2 DEPOIS de todas as outras formatações
    try:
        print('\n  Aplicando quebra de texto em H2 (encargos) e E2 (bancos)...')
        h2_cell = target_sheet['H2']
        h2_cell.alignment = Alignment(
            wrap_text=True,
            horizontal='center',
            vertical='center'
        )
        e2_cell = target_sheet['E2']
        e2_cell.alignment = Alignment(
            wrap_text=True,
            horizontal='center',
            vertical='center'
        )
        print('    ✓ H2 e E2 com quebra de texto aplicada')
    except Exception as e:
        print(f'    ⚠️  Erro ao aplicar quebra de texto: {e}')
    
    target_colors = {'F7F3DF', 'DFF0D8'}
    formulas_applied = 0
    print(f'  ✓ Aplicando fórmulas na Coluna H para cores {target_colors}...')
    try:
        for row_idx in range(1, target_sheet.max_row + 1):
            cell_h = target_sheet.cell(row=row_idx, column=8)
            bg_color = get_cell_color(cell_h)
            if bg_color in target_colors:
                formula = f'=TRUNC((G{row_idx}*(1+$G$2)),2)'
                cell_h.value = formula
                formulas_applied += 1
        print(f'  ✓ {formulas_applied} fórmulas TRUNCAR aplicadas na coluna H.')
    except Exception as e:
        print(f'    ⚠️  Erro ao aplicar fórmulas na coluna H: {e}')
    
    print('✓ Concluído')
    print('==================================================')


def apply_sintetico_step_2(workbook):
    """ETAPA 14: Aplica fórmulas em I, J e limpa/converte Coluna A (Sintético)"""
    print('\n==================================================')
    print('ETAPA 14: PROCESSANDO ORÇAMENTO SINTÉTICO (PARTE 2)')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    target_colors = {'F7F3DF', 'DFF0D8'}
    formulas_i_count = 0
    formulas_j_count = 0
    
    print(f'  ✓ Aplicando fórmulas em I & J para cores {target_colors}...')
    for row_idx in range(1, target_sheet.max_row + 1):
        try:
            cell_i = target_sheet.cell(row=row_idx, column=9)
            bg_color_i = get_cell_color(cell_i)
            if bg_color_i in target_colors:
                formula = f'=G{row_idx}*F{row_idx}'
                cell_i.value = formula
                formulas_i_count += 1
            
            cell_j = target_sheet.cell(row=row_idx, column=10)
            bg_color_j = get_cell_color(cell_j)
            if bg_color_j in target_colors:
                formula = f'=H{row_idx}*F{row_idx}'
                cell_j.value = formula
                formulas_j_count += 1
        except Exception as e:
            print(f'    ⚠️  Erro ao aplicar fórmula na linha {row_idx}: {e}')
    
    print(f'  ✓ {formulas_i_count} fórmulas aplicadas na Coluna I.')
    print(f'  ✓ {formulas_j_count} fórmulas aplicadas na Coluna J.')
    
    cells_cleaned = 0
    cells_normalized = 0
    print('\n  ✓ Normalizando Coluna A com separador ponto (a partir da linha 5)...')
    try:
        for row_idx in range(5, target_sheet.max_row + 1):
            cell_a = target_sheet.cell(row=row_idx, column=1)
            original_value = cell_a.value
            
            if original_value is None:
                continue
            
            normalized_value = normalize_sintetico_coluna_a(original_value)
            
            if isinstance(original_value, str) and original_value.replace(' ', '') != original_value:
                cells_cleaned += 1
            
            if normalized_value != original_value:
                cells_normalized += 1
            
            cell_a.value = normalized_value
            
            if normalized_value not in (None, ''):
                cell_a.number_format = '@'
    except Exception as e:
        print(f'    ⚠️  Erro ao limpar/converter Coluna A: {e}')
    
    print(f'  ✓ {cells_cleaned} células tiveram espaços removidos.')
    print(f'  ✓ {cells_normalized} células normalizadas para manter ponto (.).')
    print('✓ Concluído')
    print('==================================================')


def apply_sintetico_sum_hierarchy(workbook):
    """ETAPA 15 (Parte 1): Aplica SOMA hierárquica na planilha Sintético (sextopasso)"""
    print('\n==================================================')
    print('ETAPA 15 (Parte 1): APLICANDO SOMA HIERÁQUICA (sextopasso)')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    color_blue = 'D8ECF6'
    last_row = target_sheet.max_row
    formulas_applied = 0
    
    print(f'  ✓ Procurando linhas \'pai\' (cor {color_blue}) em {last_row} linhas...')
    for row_idx in range(1, last_row + 1):
        cell_a = target_sheet.cell(row=row_idx, column=1)
        bg_color = get_cell_color(cell_a)
        
        if bg_color == color_blue:
            sum_cells_i = []
            sum_cells_j = []
            parent_level = get_item_level_py(cell_a.value)
            
            if parent_level == -1:
                continue
            
            for child_row_idx in range(row_idx + 1, last_row + 1):
                child_cell_a = target_sheet.cell(row=child_row_idx, column=1)
                child_level = get_item_level_py(child_cell_a.value)
                
                if child_level != -1 and child_level <= parent_level:
                    break
                
                if child_level == parent_level + 1:
                    sum_cells_i.append(f'I{child_row_idx}')
                    sum_cells_j.append(f'J{child_row_idx}')
            
            if sum_cells_i:
                formula_i = f"=SUM({','.join(sum_cells_i)})"
                formula_j = f"=SUM({','.join(sum_cells_j)})"
                target_sheet.cell(row=row_idx, column=9).value = formula_i
                target_sheet.cell(row=row_idx, column=10).value = formula_j
                formulas_applied += 1
    
    print(f'  ✓ {formulas_applied} fórmulas de SOMA hierárquica aplicadas.')
    print('✓ Concluído')
    print('==================================================')


def apply_sintetico_final_totals(workbook):
    """ETAPA 15 (Parte 2): Aplica totais finais e limpa rodapé (setimopasso)"""
    print('\n==================================================')
    print('ETAPA 15 (Parte 2): APLICANDO TOTAIS FINAIS (setimopasso)')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    color_blue = 'D8ECF6'
    last_row = target_sheet.max_row
    
    level_1_cells_i = []
    level_1_cells_j = []
    total_sem_bdi_cell_j = None
    total_geral_cell_j = None
    total_do_bdi_cell_j = None
    totals_arrow_row = 0
    total_geral_row = 0
    bdi_counter = 0
    
    print('  ✓ Buscando células de Nível 1 e células de Total...')
    for row_idx in range(1, last_row + 1):
        cell_a = target_sheet.cell(row=row_idx, column=1)
        cell_i_val_labels = str(target_sheet.cell(row=row_idx, column=9).value).strip()
        cell_h_val_labels = str(target_sheet.cell(row=row_idx, column=8).value).strip()
        
        is_total_row_label = False
        
        if cell_i_val_labels == 'Total sem BDI':
            is_total_row_label = True
            bdi_counter += 1
            if bdi_counter == 2:
                total_sem_bdi_cell_j = target_sheet.cell(row=row_idx, column=10)
        elif cell_i_val_labels == 'Total Geral':
            is_total_row_label = True
            total_geral_cell_j = target_sheet.cell(row=row_idx, column=10)
            total_geral_row = row_idx
        elif cell_i_val_labels == 'Total do BDI':
            is_total_row_label = True
            total_do_bdi_cell_j = target_sheet.cell(row=row_idx, column=10)
        
        if cell_h_val_labels == 'Totais ->':
            is_total_row_label = True
            totals_arrow_row = row_idx
        
        if not is_total_row_label:
            bg_color = get_cell_color(cell_a)
            if bg_color == color_blue and get_item_level_py(cell_a.value) == 1:
                level_1_cells_i.append(f'I{row_idx}')
                level_1_cells_j.append(f'J{row_idx}')
    
    print(f'  ✓ {len(level_1_cells_i)} células de Nível 1 encontradas.')
    
    currency_format = '"$" #,##0.00'
    
    if total_sem_bdi_cell_j and level_1_cells_i:
        formula_i_total = f"=SUM({','.join(level_1_cells_i)})"
        total_sem_bdi_cell_j.value = formula_i_total
        total_sem_bdi_cell_j.number_format = currency_format
        print(f'  ✓ Fórmula inserida em {total_sem_bdi_cell_j.coordinate}: {formula_i_total}')
    
    if total_geral_cell_j and level_1_cells_j:
        formula_j_total = f"=SUM({','.join(level_1_cells_j)})"
        total_geral_cell_j.value = formula_j_total
        total_geral_cell_j.number_format = currency_format
        print(f'  ✓ Fórmula inserida em {total_geral_cell_j.coordinate}: {formula_j_total}')
    
    if total_do_bdi_cell_j and total_geral_cell_j and total_sem_bdi_cell_j:
        formula_bdi = f'={total_geral_cell_j.coordinate}-{total_sem_bdi_cell_j.coordinate}'
        total_do_bdi_cell_j.value = formula_bdi
        total_do_bdi_cell_j.number_format = currency_format
        print(f'  ✓ Fórmula inserida em {total_do_bdi_cell_j.coordinate}: {formula_bdi}')
    
    if totals_arrow_row > 0:
        for col_idx in range(1, target_sheet.max_column + 1):
            clear_cell(target_sheet.cell(row=totals_arrow_row, column=col_idx))
        print(f'  ✓ Conteúdo da linha {totals_arrow_row} (\'Totais ->\') limpo.')
    
    if total_geral_row > 0 and total_geral_row < target_sheet.max_row:
        rows_to_clear_count = 0
        for row_to_clear in range(total_geral_row + 1, target_sheet.max_row + 1):
            for col_to_clear in range(1, target_sheet.max_column + 1):
                clear_cell(target_sheet.cell(row=row_to_clear, column=col_to_clear))
            rows_to_clear_count += 1
        print(f'  ✓ Conteúdo de {rows_to_clear_count} linhas abaixo de \'Total Geral\' limpo.')
    
    print('✓ Concluído')
    print('==================================================')


def apply_sintetico_step_3(workbook):
    """ETAPA 16: Move totais para G/H e aplica fórmulas de percentual em K"""
    print('\n==================================================')
    print('ETAPA 16: MOVENDO TOTAIS E CALCULANDO PERCENTUAIS')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    target_labels = {'Total Geral', 'Total sem BDI', 'Total do BDI'}
    rows_moved = []
    total_locations = {}
    
    print('  ✓ Procurando rótulos de total na Coluna I para mover...')
    for row_idx in range(1, target_sheet.max_row + 1):
        cell_i = target_sheet.cell(row=row_idx, column=9)
        cell_i_val = str(cell_i.value).strip()
        
        if cell_i_val in target_labels:
            cell_j = target_sheet.cell(row=row_idx, column=10)
            cell_g = target_sheet.cell(row=row_idx, column=7)
            cell_h = target_sheet.cell(row=row_idx, column=8)
            
            copy_cell_style_and_value(cell_i, cell_g)
            copy_cell_style_and_value(cell_j, cell_h)
            
            clear_cell(cell_i)
            clear_cell(cell_j)
            
            rows_moved.append(row_idx)
            total_locations[cell_i_val] = row_idx
            print(f'    ✓ Rótulo \'{cell_i_val}\' movido da linha {row_idx} (I/J -> G/H)')
    
    if 'Total do BDI' in total_locations and 'Total Geral' in total_locations and 'Total sem BDI' in total_locations:
        try:
            bdi_row = total_locations['Total do BDI']
            geral_row = total_locations['Total Geral']
            sem_bdi_row = total_locations['Total sem BDI']
            
            cell_h_bdi = target_sheet.cell(row=bdi_row, column=8)
            geral_coord = target_sheet.cell(row=geral_row, column=8).coordinate
            sem_bdi_coord = target_sheet.cell(row=sem_bdi_row, column=8).coordinate
            
            formula_bdi = f'={geral_coord}-{sem_bdi_coord}'
            cell_h_bdi.value = formula_bdi
            cell_h_bdi.number_format = '"$" #,##0.00'
            print(f'  ✓ Fórmula do \'Total do BDI\' em {cell_h_bdi.coordinate} atualizada')
        except Exception as e:
            print(f'    ⚠️  Falha ao atualizar fórmula do BDI: {e}')
    
    if 'Total Geral' not in total_locations:
        print('  ⚠️  \'Total Geral\' não encontrado.')
        print('==================================================')
        return
    
    total_geral_row = total_locations['Total Geral']
    total_geral_H_ref = f'$H${total_geral_row}'
    target_colors = {'DFF0D8', 'D8ECF6', 'F7F3DF'}
    formulas_K_applied = 0
    
    print(f'  ✓ Aplicando fórmulas de percentual na Coluna K...')
    for row_idx in range(1, target_sheet.max_row + 1):
        if row_idx in rows_moved:
            continue
        
        cell_to_check_J = target_sheet.cell(row=row_idx, column=10)
        cell_to_check_H = target_sheet.cell(row=row_idx, column=8)
        bg_color = get_cell_color(cell_to_check_J) or get_cell_color(cell_to_check_H)
        
        if bg_color in target_colors:
            cell_K = target_sheet.cell(row=row_idx, column=11)
            formula_k = f'=J{row_idx}/{total_geral_H_ref}'
            cell_K.value = formula_k
            cell_K.number_format = '0.00%'
            formulas_K_applied += 1
    
    print(f'  ✓ {formulas_K_applied} fórmulas de percentual aplicadas na Coluna K.')
    print('✓ Concluído')
    print('==================================================')


def apply_sintetico_step_4(workbook, problematic_codes):
    """ETAPA 17: Preenche valores na Coluna G do Sintético (oitavopasso)"""
    print('\n==================================================')
    print('ETAPA 17: PREENCHENDO COLUNA G DO SINTÉTICO (oitavopasso)')
    print('==================================================')
    wsOS = None
    wsCPUs = None
    wsABC = None
    
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            wsOS = sheet
        elif 'CPU' in sheet.title or 'Compos' in sheet.title:
            wsCPUs = sheet
        elif 'Curva ABC de Insumos' in sheet.title:
            wsABC = sheet
    
    if wsOS is None or wsCPUs is None or wsABC is None:
        print('  ⚠️  Uma das planilhas não foi encontrada.')
        print('==================================================')
        return
    
    print(f'  ✓ Planilhas encontradas: {wsOS.title}, {wsCPUs.title}, {wsABC.title}')
    
    color_green = 'DFF0D8'
    color_yellow = 'F7F3DF'
    cpus_sheet_name = f'\'{wsCPUs.title}\''
    abc_sheet_name = f'\'{wsABC.title}\''
    
    lastRowOS = wsOS.max_row
    formulas_green = 0
    formulas_yellow_normal = 0
    formulas_yellow_problem = 0
    
    print(f'  ✓ Varrendo {lastRowOS} linhas da planilha \'{wsOS.title}\'...')
    for row_idx in range(1, lastRowOS + 1):
        try:
            cell_g = wsOS.cell(row=row_idx, column=7)
            cellColor = get_cell_color(cell_g)
            
            if cellColor == color_green:
                cell_b_ref = f'B{row_idx}'
                formula_en = f'=IFERROR(VLOOKUP({cell_b_ref},{cpus_sheet_name}!$B:$H,7,FALSE),0)'
                cell_g.value = formula_en
                formulas_green += 1
            elif cellColor == color_yellow:
                code = str(wsOS.cell(row=row_idx, column=2).value).strip()
                
                if code in problematic_codes:
                    cell_d_ref = f'D{row_idx}'
                    formula_en = f'=IFERROR(INDEX({abc_sheet_name}!$G:$G,MATCH({cell_d_ref},{abc_sheet_name}!$C:$C,0)),0)'
                    cell_g.value = formula_en
                    formulas_yellow_problem += 1
                else:
                    cell_b_ref = f'B{row_idx}'
                    formula_en = f'=IFERROR(VLOOKUP({cell_b_ref},{abc_sheet_name}!$A:$G,7,FALSE),0)'
                    cell_g.value = formula_en
                    formulas_yellow_normal += 1
        except Exception as e:
            print(f'    ⚠️  Erro ao processar linha {row_idx}: {e}')
    
    print(f'  ✓ {formulas_green} fórmulas para \'Composições\' (Verde).')
    print(f'  ✓ {formulas_yellow_normal} fórmulas para \'Insumos Normais\' (Amarelo).')
    print(f'  ✓ {formulas_yellow_problem} fórmulas para \'Insumos Problemáticos\' (Amarelo).')
    print('✓ Concluído')
    print('==================================================')


def apply_nono_passo_sintetico(workbook):
    """ETAPA 21: Aplica o Nono Passo (nonopasso) - Macro 78"""
    print('\n==================================================')
    print('ETAPA 21: APLICANDO NONO PASSO (nonopasso - Macro 78)')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title or 'Orçamento Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    last_row = target_sheet.max_row
    
    parent_addresses_i = []
    parent_addresses_j = []
    total_sem_bdi_cell_h = None
    total_geral_cell_h = None
    
    print('  ✓ Procurando itens de nível 1 (começando da linha 5)...')
    items_nivel_1_encontrados = 0
    
    for row_idx in range(5, last_row + 1):
        cell_a = target_sheet.cell(row=row_idx, column=1)
        if get_item_level_py(cell_a.value) == 1:
            cell_i_address = f'I{row_idx}'
            cell_j_address = f'J{row_idx}'
            parent_addresses_i.append(cell_i_address)
            parent_addresses_j.append(cell_j_address)
            items_nivel_1_encontrados += 1
    
    print(f'  ✓ Total de itens de nível 1 encontrados: {items_nivel_1_encontrados}')
    
    print('  ✓ Procurando células \'Total sem BDI\' e \'Total Geral\' na coluna G...')
    for row_idx in range(1, last_row + 1):
        cell_g = target_sheet.cell(row=row_idx, column=7)
        cell_g_value = str(cell_g.value).strip() if cell_g.value else ''
        
        if cell_g_value == 'Total sem BDI':
            total_sem_bdi_cell_h = target_sheet.cell(row=row_idx, column=8)
        elif cell_g_value == 'Total Geral':
            total_geral_cell_h = target_sheet.cell(row=row_idx, column=8)
    
    currency_format = '"$" #,##0.00'
    
    if total_sem_bdi_cell_h and len(parent_addresses_i) > 0:
        formula_i_total = f"=SUM({','.join(parent_addresses_i)})"
        total_sem_bdi_cell_h.value = formula_i_total
        total_sem_bdi_cell_h.number_format = currency_format
        print(f'  ✓ Fórmula inserida em {total_sem_bdi_cell_h.coordinate}')
    
    if total_geral_cell_h and len(parent_addresses_j) > 0:
        formula_j_total = f"=SUM({','.join(parent_addresses_j)})"
        total_geral_cell_h.value = formula_j_total
        total_geral_cell_h.number_format = currency_format
        print(f'  ✓ Fórmula inserida em {total_geral_cell_h.coordinate}')
    
    print('✓ Nono Passo (nonopasso) concluído!')
    print('==================================================')


def apply_sintetico_formatting(workbook):
    """ETAPA 18: Aplica formatação final na planilha Sintético (bordas e título)"""
    print('\n==================================================')
    print('ETAPA 18: APLICANDO FORMATAÇÃO FINAL NO SINTÉTICO')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    
    last_row = target_sheet.max_row
    last_col = 11
    
    thin_border_side = Side(style='thin')
    all_borders = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    no_border = Border()
    
    last_colored_row = 0
    total_keywords = ['Total sem BDI', 'Total do BDI', 'Total Geral', 'Totais']
    
    for row_idx in range(last_row, 0, -1):
        is_total_row = False
        for col_idx in range(1, last_col + 1):
            cell_value = str(target_sheet.cell(row=row_idx, column=col_idx).value or '')
            if any(keyword in cell_value for keyword in total_keywords):
                is_total_row = True
                break
        
        if is_total_row:
            continue
        
        for col_idx in range(1, last_col + 1):
            cell = target_sheet.cell(row=row_idx, column=col_idx)
            cell_color = get_cell_color(cell)
            if cell_color is not None:
                last_colored_row = row_idx
                break
        
        if last_colored_row > 0:
            break
    
    if last_colored_row == 0:
        last_colored_row = 1
    
    print(f'✓ Última linha com cor detectada: {last_colored_row}')
    print(f'  ✓ Aplicando bordas em A1:K{last_colored_row}...')
    
    for row_idx in range(1, last_colored_row + 1):
        for col_idx in range(1, last_col + 1):
            target_sheet.cell(row=row_idx, column=col_idx).border = all_borders
    
    for row_idx in range(last_colored_row + 1, last_row + 1):
        for col_idx in range(1, last_col + 1):
            target_sheet.cell(row=row_idx, column=col_idx).border = no_border
    
    arial_bold_font = Font(name='Arial', bold=True)
    for row_num in [1, 2, 4]:
        for col_idx in range(1, last_col + 1):
            cell = target_sheet.cell(row=row_num, column=col_idx)
            cell.font = arial_bold_font
    
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    white_font = Font(color='FFFFFF')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    try:
        merge_range = 'A3:K3'
        target_sheet.merge_cells(merge_range)
        cell_a3 = target_sheet['A3']
        cell_a3.value = 'Orçamento Sintético'
        cell_a3.fill = blue_fill
        cell_a3.font = white_font
        cell_a3.alignment = center_alignment
    except Exception as e:
        print(f'    ⚠️  Erro ao formatar Linha 3: {e}')
    
    print('✓ Concluído')
    print('==================================================')


def apply_sintetico_final_merges(workbook):
    """ETAPA 19: Aplica mesclagem final de cabeçalhos no Sintético"""
    print('\n==================================================')
    print('ETAPA 19: APLICANDO MESCLAGEM FINAL NO SINTÉTICO')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    if target_sheet is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet encontrada: {target_sheet.title}')
    center_middle_alignment = Alignment(horizontal='center', vertical='center')
    
    merges_to_apply = ['A1:C2', 'E1:F1', 'E2:F2', 'H2:J2', 'H1:J1']
    
    for merge_range in merges_to_apply:
        try:
            target_sheet.merge_cells(merge_range)
            print(f'  ✓ Células {merge_range} mescladas.')
        except Exception as e:
            if 'already merged' not in str(e):
                print(f'    ⚠️  Erro ao mesclar {merge_range}: {e}')
    
    try:
        target_sheet.column_dimensions['H'].width = 15
        target_sheet.column_dimensions['I'].width = 15
        target_sheet.column_dimensions['J'].width = 15
    except Exception as e:
        print(f'    ⚠️  Erro ao ajustar larguras: {e}')
    
    try:
        for row in target_sheet.iter_rows(min_row=1, max_row=2, min_col=5, max_col=11):
            for cell in row:
                cell.alignment = center_middle_alignment
    except Exception as e:
        print(f'    ⚠️  Erro ao aplicar alinhamento: {e}')
    
    print('✓ Concluído')
    print('==================================================')


def apply_abc_final_formatting(workbook):
    """ETAPA 20: Aplica formatação final na Curva ABC (mesclagem e fórmulas de link)"""
    print('\n==================================================')
    print('ETAPA 20: APLICANDO FORMATAÇÃO FINAL NA CURVA ABC')
    print('==================================================')
    wsABC = None
    wsOS = None
    
    for sheet in workbook.worksheets:
        if 'Curva ABC de Insumos' in sheet.title:
            wsABC = sheet
        elif 'Sintético' in sheet.title:
            wsOS = sheet
    
    if wsABC is None:
        print('  ⚠️  Sheet \'Curva ABC de Insumos\' não encontrada')
        print('==================================================')
        return
    if wsOS is None:
        print('  ⚠️  Sheet \'Orçamento Sintético\' não encontrada')
        print('==================================================')
        return
    
    print(f'  ✓ Sheet \'Curva ABC\' encontrada: {wsABC.title}')
    print(f'  ✓ Sheet \'Sintético\' encontrada: {wsOS.title}')
    
    merges_to_apply = ['A1:B2', 'D1:E1', 'D2:E2', 'G1:H1', 'G2:H2']
    
    for merge_range in merges_to_apply:
        try:
            wsABC.merge_cells(merge_range)
        except Exception as e:
            if 'already merged' not in str(e):
                print(f'    ⚠️  Erro ao mesclar {merge_range}: {e}')
    
    total_geral_row = -1
    try:
        for row_idx in range(1, wsOS.max_row + 1):
            cell_g_val = str(wsOS.cell(row=row_idx, column=7).value).strip()
            if cell_g_val == 'Total Geral':
                total_geral_row = row_idx
                break
        
        if total_geral_row != -1:
            formula_i2 = f'=\'{wsOS.title}\'!H{total_geral_row}'
            wsABC['I2'].value = formula_i2
        
        formula_f2 = f'=\'{wsOS.title}\'!G2'
        wsABC['F2'].value = formula_f2
    except Exception as e:
        print(f'    ⚠️  Erro ao criar fórmulas: {e}')
    
    print('✓ Concluído')
    print('==================================================')


def apply_aesthetic_adjustments(workbook):
    """ETAPA ESTÉTICA: Ajustes visuais finais nas planilhas"""
    print('\n==================================================')
    print('ETAPA ESTÉTICA: APLICANDO AJUSTES VISUAIS FINAIS')
    print('==================================================')
    sintetico_sheet = None
    abc_sheet = None
    cpus_sheet = None
    
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            sintetico_sheet = sheet
        elif 'Curva ABC' in sheet.title:
            abc_sheet = sheet
        elif 'CPU' in sheet.title or 'Compos' in sheet.title:
            cpus_sheet = sheet
    
    if sintetico_sheet and abc_sheet:
        try:
            source_cell = sintetico_sheet['E2']
            target_cell = abc_sheet['D2']
            copy_cell_style_and_value(source_cell, target_cell)
            print(f'  ✓ E2 (Sintético) copiado para D2 (ABC)')
        except Exception as e:
            print(f'  ⚠️  Erro ao copiar E2 para D2: {e}')
    
    new_blue = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    
    if sintetico_sheet:
        try:
            for col_idx in range(1, sintetico_sheet.max_column + 1):
                cell = sintetico_sheet.cell(3, col_idx)
                bg_color = get_cell_color(cell)
                if bg_color and bg_color in ['0000FF', '00FF']:
                    cell.fill = new_blue
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            print('  ✓ Cor azul da linha 3 alterada para #366092 (Sintético)')
        except Exception as e:
            print(f'  ⚠️  Erro: {e}')
    
    if abc_sheet:
        try:
            for col_idx in range(1, abc_sheet.max_column + 1):
                cell = abc_sheet.cell(3, col_idx)
                bg_color = get_cell_color(cell)
                if bg_color and bg_color in ['0000FF', '00FF']:
                    cell.fill = new_blue
                    cell.font = white_font
        except Exception as e:
            print(f'  ⚠️  Erro: {e}')
    
    if cpus_sheet:
        try:
            for row_idx in [1, 2]:
                for col_idx in range(1, cpus_sheet.max_column + 1):
                    cell = cpus_sheet.cell(row_idx, col_idx)
                    bg_color = get_cell_color(cell)
                    if bg_color and bg_color in ['0000FF', '00FF']:
                        cell.fill = new_blue
                        cell.font = white_font
        except Exception as e:
            print(f'  ⚠️  Erro: {e}')
    
    print('✓ Ajustes estéticos aplicados!')
    print('==================================================')


def finalize_sintetico_headers(workbook):
    """ETAPA FINAL: Garantir que os cabeçalhos G4:J4 estejam corretos"""
    print('\n==================================================')
    print('ETAPA FINAL: GARANTINDO CABEÇALHOS E FORMATAÇÕES FINAIS')
    print('==================================================')
    sintetico_sheet = None
    abc_sheet = None
    
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            sintetico_sheet = sheet
        elif 'Curva ABC' in sheet.title or 'Curva ABC de Insumos' in sheet.title:
            abc_sheet = sheet
    
    if sintetico_sheet:
        try:
            sintetico_sheet['E4'].value = 'Bancos'
            sintetico_sheet['F4'].value = 'B.D.I.'
            sintetico_sheet['G4'].value = 'Valor Unit'
            sintetico_sheet['H4'].value = 'Valor Unit com BDI'
            sintetico_sheet['I4'].value = 'Total sem BDI'
            sintetico_sheet['J4'].value = 'Total com BDI'
            sintetico_sheet['K4'].value = 'Peso (%)'
            print('  ✓ Cabeçalhos E4:K4 definidos')
        except Exception as e:
            print(f'  ⚠️  Erro: {e}')
        
        # Aplicar formatação contábil R$ nas colunas G, H, I (exceto G2)
        try:
            print('\n  ✓ Aplicando formatação contábil R$ nas colunas G, H, I...')
            accounting_format = '_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'
            formatted_cells = 0
            
            for row_idx in range(1, sintetico_sheet.max_row + 1):
                for col_idx in [7, 8, 9]:  # Colunas G, H, I
                    # Exceção: não formatar G2 (contém porcentagem de BDI)
                    if row_idx == 2 and col_idx == 7:
                        continue
                    
                    cell = sintetico_sheet.cell(row_idx, col_idx)
                    if cell.value is not None:
                        # Aplicar formato apenas se for número ou fórmula
                        if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                            cell.number_format = accounting_format
                            formatted_cells += 1
            
            print(f'  ✓ Formatação contábil R$ aplicada em {formatted_cells} células (G, H, I - exceto G2)')
        except Exception as e:
            print(f'  ⚠️  Erro ao aplicar formatação contábil: {e}')
    
    if abc_sheet:
        try:
            abc_sheet['G1'].value = 'Encargos Sociais'
            abc_sheet['K4'].value = 'Desconto'
            print('  ✓ Headers ABC atualizados')
        except Exception as e:
            print(f'  ⚠️  Erro: {e}')
        
        # Copiar H2 do Sintético para G2 da Curva ABC
        if sintetico_sheet:
            try:
                print('\n  ✓ Copiando H2 do Sintético para G2 da Curva ABC...')
                source_h2 = sintetico_sheet['H2']
                target_g2 = abc_sheet['G2']
                
                # Copiar valor e formatação
                copy_cell_style_and_value(source_h2, target_g2)
                
                print(f'  ✓ H2 do Sintético copiado para G2 da ABC: \'{source_h2.value}\'')
            except Exception as e:
                print(f'  ⚠️  Erro ao copiar H2 para G2: {e}')
    
    # APLICAR QUEBRA DE TEXTO FINAL (última etapa para garantir que não seja sobrescrita)
    print('\n  ✓ Aplicando quebra de texto FINAL em todas as células necessárias...')
    
    if sintetico_sheet:
        try:
            # H2 (Encargos Sociais)
            h2_cell = sintetico_sheet['H2']
            h2_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )
            # E2 (Bancos)
            e2_cell = sintetico_sheet['E2']
            e2_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )
            print('    ✓ Orçamento Sintético: H2 e E2 com quebra de texto')
        except Exception as e:
            print(f'    ⚠️  Erro em Sintético: {e}')
    
    if abc_sheet:
        try:
            # G2 (Encargos Sociais)
            g2_cell = abc_sheet['G2']
            g2_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )
            # E2 (Bancos)
            e2_abc_cell = abc_sheet['E2']
            e2_abc_cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center'
            )
            print('    ✓ Curva ABC: G2 e E2 com quebra de texto')
            
            # Garantir formatação com separador de milhar na coluna F
            print('    ✓ Aplicando separador de milhar na coluna F da Curva ABC...')
            for row_idx in range(1, abc_sheet.max_row + 1):
                cell = abc_sheet.cell(row_idx, 6)  # Coluna F
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            print('    ✓ Formatação de milhar aplicada na coluna F')
        except Exception as e:
            print(f'    ⚠️  Erro em ABC: {e}')
    
    print('✓ Finalizações aplicadas!')
    print('==================================================')


def enforce_sintetico_coluna_a_dot_rule(workbook):
    """ETAPA COMPLEMENTAR: Força separador de ponto na Coluna A do Sintético"""
    print('\n==================================================')
    print('COMPLEMENTAR: FORÇANDO PONTOS NA COLUNA A (SINTÉTICO)')
    print('==================================================')
    target_sheet = None
    for sheet in workbook.worksheets:
        if 'Sintético' in sheet.title:
            target_sheet = sheet
            break
    
    if target_sheet is None:
        print('  ⚠️  Sintético não encontrado')
        print('==================================================')
        return
    
    print(f'  ✓ Processando: {target_sheet.title}')
    converted = 0
    
    for row_idx in range(5, target_sheet.max_row + 1):
        cell_a = target_sheet.cell(row=row_idx, column=1)
        if cell_a.value:
            normalized = normalize_sintetico_coluna_a(cell_a.value)
            if normalized != cell_a.value:
                cell_a.value = normalized
                cell_a.number_format = '@'
                converted += 1
    
    print(f'  ✓ {converted} células normalizadas')
    print('==================================================')


def _find_sheet_by_contains(workbook: Workbook, token: str):
    """Encontra uma sheet que contém o token no título"""
    token_lower = token.lower()
    for sheet in workbook.worksheets:
        if token_lower in sheet.title.lower():
            return sheet
    return None


def validate_input_file(file_path: Path) -> None:
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
    if not file_path.is_file():
        raise ValueError(f"O caminho não é um arquivo: {file_path}")
    if file_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        exts = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Formato inválido para '{file_path.name}'. Use: {exts}")


def make_unique_sheet_title(workbook: Workbook, desired_title: str) -> str:
    base_title = (desired_title or "Planilha").strip()[:31] or "Planilha"
    existing = {sheet.title for sheet in workbook.worksheets}

    if base_title not in existing:
        return base_title

    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base_title[:31 - len(suffix)]}{suffix}"
        if candidate not in existing:
            return candidate
        index += 1


def copy_sheet_with_formatting(source: Worksheet, target: Worksheet) -> None:
    for col_letter, col_dimension in source.column_dimensions.items():
        target.column_dimensions[col_letter].width = col_dimension.width
        target.column_dimensions[col_letter].hidden = col_dimension.hidden

    for row_idx, row_dimension in source.row_dimensions.items():
        target.row_dimensions[row_idx].height = row_dimension.height
        target.row_dimensions[row_idx].hidden = row_dimension.hidden

    for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=source.max_column):
        for source_cell in row:
            target_cell = target.cell(row=source_cell.row, column=source_cell.column)
            target_cell.value = source_cell.value

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))

    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)


def unify_spreadsheets(input_files: list[Path], output_path: Path, progress_callback=None) -> Path:
    """Unifica planilhas com callback opcional para progresso.
    
    Args:
        input_files: Lista de arquivos de entrada
        output_path: Caminho do arquivo de saída
        progress_callback: Função que recebe (percentual, mensagem) para atualizar progresso
    """
    def update_progress(percent: float, message: str):
        if progress_callback:
            progress_callback(percent, message)
    
    update_progress(5, "Criando arquivo consolidado...")
    consolidated = Workbook()
    consolidated.remove(consolidated.active)

    # Processamento dos arquivos: 5% a 35%
    for index, input_file in enumerate(input_files):
        progress_percent = 5 + (index * 10)
        update_progress(progress_percent, f"Carregando arquivo {index + 1} de 3: {input_file.name}")
        
        source_wb = load_workbook(input_file, data_only=False)
        try:
            if not source_wb.worksheets:
                raise ValueError(f"Arquivo sem planilhas: {input_file}")

            source_sheet = source_wb.worksheets[0]
            target_name = TARGET_SHEET_NAMES[index] if index < len(TARGET_SHEET_NAMES) else source_sheet.title
            new_title = make_unique_sheet_title(consolidated, target_name)
            new_sheet = consolidated.create_sheet(title=new_title)
            
            update_progress(progress_percent + 5, f"Copiando dados do arquivo {index + 1}...")
            copy_sheet_with_formatting(source_sheet, new_sheet)
        finally:
            source_wb.close()

    # Aplicar processamento completo
    update_progress(40, "Aplicando processamento de planilhas...")
    apply_legacy_processing(consolidated, progress_callback)
    
    update_progress(95, "Salvando arquivo consolidado...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    consolidated.save(output_path)
    consolidated.close()
    
    update_progress(100, "Concluído!")
    return output_path


def _find_sheet_by_contains(workbook: Workbook, token: str):
    token_lower = token.lower()
    for sheet in workbook.worksheets:
        if token_lower in sheet.title.lower():
            return sheet
    return None


def apply_legacy_processing(workbook: Workbook, progress_callback=None) -> None:
    """Aplica processamento completo às planilhas consolidadas.
    
    Esta função agora usa as funções de processamento locais, não mais um módulo externo.
    """
    def update_progress(percent: float, message: str):
        if progress_callback:
            progress_callback(percent, message)
    
    # ETAPA 2: Desmerge
    update_progress(40, "Removendo células mescladas...")
    unmerge_all_cells(workbook)
    
    # ETAPAS 3-6: Curva ABC
    update_progress(45, "Processando Curva ABC de Insumos...")
    process_curva_abc_sheet(workbook)
    
    update_progress(47, "Aplicando fórmulas na Curva ABC (colunas G e H)...")
    apply_formulas_to_curva_abc(workbook)
    
    update_progress(49, "Aplicando soma e fórmulas na Curva ABC (H, I, K)...")
    apply_sum_and_formulas_to_curva_abc(workbook)
    
    update_progress(51, "Aplicando soma coluna I e deletando linhas...")
    apply_sum_column_i_and_delete_below(workbook)
    
    # ETAPAS 7-9: CPUs (primeira parte)
    update_progress(53, "Processando planilha CPUs...")
    process_cpus_sheet(workbook)
    
    update_progress(55, "Limpando linhas específicas CPUs...")
    clear_specific_rows_cpus(workbook)
    
    update_progress(57, "Aplicando fórmulas na planilha CPUs...")
    apply_formulas_cpus_sheet(workbook)
    
    update_progress(59, "Aplicando referências para composições auxiliares...")
    apply_third_step_cpus_sheet(workbook)
    
    # ETAPA 10: Identificar códigos problemáticos e aplicar VLOOKUP/INDEX-MATCH
    update_progress(60, "Identificando códigos problemáticos...")
    problematic_codes = set()
    abc_sheet = _find_sheet_by_contains(workbook, "Curva ABC de Insumos")
    if abc_sheet is not None:
        problematic_codes = _find_problematic_abc_codes(abc_sheet)
    
    update_progress(62, "Aplicando busca de insumos na Curva ABC...")
    apply_fourth_step_cpus_sheet(workbook, problematic_codes)
    
    # ETAPAS 11-12: CPUs (segunda parte)
    update_progress(64, "Limpando linhas verdes com coluna D vazia...")
    apply_fifth_step_cpus_sheet(workbook)
    
    update_progress(66, "Formatando cabeçalhos e limpando final da CPU...")
    apply_sixth_step_cpus_sheet(workbook)
    
    # ETAPAS 13-14: Sintético (primeira parte)
    update_progress(68, "Processando Orçamento Sintético (parte 1)...")
    process_sintetico_sheet(workbook)
    
    update_progress(70, "Processando Orçamento Sintético (parte 2)...")
    apply_sintetico_step_2(workbook)
    
    # ETAPA 17: Preencher coluna G do Sintético (precisa vir depois da linha 5 ser deletada)
    update_progress(72, "Preenchendo coluna G do Sintético...")
    apply_sintetico_step_4(workbook, problematic_codes)
    
    # ETAPAS 15-16: Sintético (somas hierárquicas e totais)
    update_progress(74, "Aplicando soma hierárquica no Sintético...")
    apply_sintetico_sum_hierarchy(workbook)
    
    update_progress(76, "Aplicando totais finais no Sintético...")
    apply_sintetico_final_totals(workbook)
    
    update_progress(78, "Movendo totais e calculando percentuais...")
    apply_sintetico_step_3(workbook)
    
    # ETAPA 21: Nono passo
    update_progress(80, "Aplicando nono passo (nonopasso)...")
    apply_nono_passo_sintetico(workbook)
    
    # ETAPAS 18-20: Formatações finais
    update_progress(82, "Aplicando formatação final no Sintético...")
    apply_sintetico_formatting(workbook)
    
    update_progress(84, "Aplicando mesclagem final no Sintético...")
    apply_sintetico_final_merges(workbook)
    
    update_progress(86, "Aplicando formatação final na Curva ABC...")
    apply_abc_final_formatting(workbook)
    
    # ETAPAS ESTÉTICAS E FINAIS
    update_progress(88, "Aplicando ajustes estéticos...")
    apply_aesthetic_adjustments(workbook)
    
    update_progress(90, "Finalizando cabeçalhos...")
    finalize_sintetico_headers(workbook)
    
    update_progress(92, "Forçando separador de ponto na coluna A...")
    enforce_sintetico_coluna_a_dot_rule(workbook)
    
    update_progress(95, "Processamento completo!")



def build_output_path(custom_output: str | None) -> Path:
    if custom_output:
        output = Path(custom_output).expanduser().resolve()
        if output.suffix.lower() not in SUPPORTED_EXTENSIONS:
            output = output.with_suffix(".xlsx")
        return output

    downloads = Path.home() / "Downloads"
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return downloads / f"Planilha_Consolidada_{timestamp}.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Unifica até 3 planilhas Excel em um único arquivo (sem servidor web)."
    )
    parser.add_argument(
        "files",
        nargs="*",
        help="Caminhos dos 3 arquivos de entrada (xlsx/xlsm/xltx/xltm).",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Caminho completo do arquivo de saída. Se omitido, salva em Downloads.",
    )
    parser.add_argument(
        "--cli",
        action="store_true",
        help="Força execução em modo linha de comando.",
    )
    return parser.parse_args()


def prompt_for_files_if_needed(cli_files: list[str]) -> list[Path]:
    if cli_files:
        files = [Path(item).expanduser().resolve() for item in cli_files]
    else:
        print("Informe os 3 arquivos de entrada:")
        files = []
        for index in range(1, 4):
            file_input = input(f"Arquivo {index}: ").strip().strip('"')
            files.append(Path(file_input).expanduser().resolve())

    if len(files) != 3:
        raise ValueError("É necessário informar exatamente 3 arquivos de entrada.")

    return files


class TextRedirector:
    """Redireciona stdout para um widget Text do Tkinter"""
    def __init__(self, text_widget: tk.Text):
        self.text_widget = text_widget

    def write(self, string: str):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)
        self.text_widget.update_idletasks()

    def flush(self):
        pass


class UnificadorApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Unificador de Planilhas")
        self.root.geometry("1180x820")
        self.root.minsize(1080, 720)
        self.root.configure(bg="#08587A")
        self.root.state('zoomed')  # Abrir maximizado
        
        # Definir ícone da janela (funciona tanto em desenvolvimento quanto como executável)
        try:
            # Quando empacotado com PyInstaller, os arquivos estão em sys._MEIPASS
            if getattr(sys, 'frozen', False):
                # Executável PyInstaller
                base_path = Path(sys._MEIPASS)
            else:
                # Desenvolvimento
                base_path = Path(__file__).parent
            
            icon_path = base_path / "icon.ico"
            if icon_path.exists():
                self.root.iconbitmap(str(icon_path))
        except Exception as e:
            print(f"Aviso: Não foi possível carregar o ícone: {e}")
            pass  # Se não conseguir carregar o ícone, continua sem ele

        self.colors = {
            "bg": "#08587A",
            "card_bg": "#F0F1F3",
            "card_border": "#1A7FA8",
            "title": "#EAC248",
            "subtitle": "#E8F2F7",
            "text_primary": "#0C4560",
            "text_secondary": "#606B74",
            "button_bg": "#1188B2",
            "button_text": "#FFFFFF",
            "button_disabled": "#BEBFC1",
            "ok": "#1D7F43",
            "error": "#A11D2A",
        }

        self.file_vars = [tk.StringVar(), tk.StringVar(), tk.StringVar()]
        self.file_name_vars = [
            tk.StringVar(value="Nenhum arquivo selecionado"),
            tk.StringVar(value="Nenhum arquivo selecionado"),
            tk.StringVar(value="Nenhum arquivo selecionado"),
        ]
        self.status_var = tk.StringVar(value="Selecione as 3 planilhas para começar.")

        self._build_ui()
        self._update_generate_button_state()

    def _build_ui(self) -> None:
        header = tk.Frame(self.root, bg=self.colors["bg"])
        header.pack(fill="x", pady=(16, 10))

        title = tk.Label(
            header,
            text="Unificador de Planilhas",
            font=("Segoe UI", 34, "bold"),
            fg=self.colors["title"],
            bg=self.colors["bg"],
        )
        title.pack()

        subtitle = tk.Label(
            header,
            text="Sistema de Consolidação Automática",
            font=("Segoe UI", 20),
            fg=self.colors["subtitle"],
            bg=self.colors["bg"],
        )
        subtitle.pack(pady=(0, 10))

        container = tk.Frame(
            self.root,
            bg=self.colors["card_bg"],
            highlightbackground="#D6DADF",
            highlightthickness=1,
            bd=0,
            padx=18,
            pady=18,
        )
        container.pack(fill="both", expand=True, padx=26, pady=(4, 26))

        cards = tk.Frame(container, bg=self.colors["card_bg"])
        cards.pack(fill="x", pady=(8, 14))

        self._build_file_card(cards, 0, "📊", "Planilha 1", "Composições com Preço Unitário")
        self._build_file_card(cards, 1, "📈", "Planilha 2", "Curva ABC de Insumos")
        self._build_file_card(cards, 2, "💼", "Planilha 3", "Sintético com Valor da Mão de Obra")

        divider = tk.Frame(container, bg="#D0D2D6", height=2)
        divider.pack(fill="x", padx=22, pady=(6, 18))

        action_frame = tk.Frame(container, bg=self.colors["card_bg"])
        action_frame.pack(fill="x", pady=(4, 8))

        self.generate_button = tk.Button(
            action_frame,
            text="Gerar Planilha Final",
            command=self.generate,
            font=("Segoe UI", 18, "bold"),
            bg=self.colors["button_bg"],
            fg=self.colors["button_text"],
            activebackground="#0E7CA2",
            activeforeground=self.colors["button_text"],
            relief="flat",
            bd=0,
            padx=40,
            pady=14,
            cursor="hand2",
            disabledforeground="#5D5D5D",
        )
        self.generate_button.pack(anchor="center")

        status_frame = tk.Frame(container, bg=self.colors["card_bg"])
        status_frame.pack(fill="x", pady=(14, 0), padx=24)

        self.status_label = tk.Label(
            status_frame,
            textvariable=self.status_var,
            wraplength=980,
            justify="left",
            anchor="w",
            font=("Segoe UI", 12),
            fg=self.colors["text_primary"],
            bg=self.colors["card_bg"],
        )
        self.status_label.pack(fill="x")

        # Frame de progresso (inicialmente oculto)
        self.progress_frame = tk.Frame(
            container, 
            bg="#FFFFFF",  # Fundo branco para destacar
            highlightbackground="#1188B2",
            highlightthickness=2,
            relief="ridge"
        )
        
        progress_title = tk.Label(
            self.progress_frame,
            text="⏳ Processando...",
            font=("Segoe UI", 16, "bold"),
            fg=self.colors["text_primary"],
            bg="#FFFFFF",
        )
        progress_title.pack(pady=(20, 10))

        # Canvas para o spinner animado
        spinner_container = tk.Frame(self.progress_frame, bg="#FFFFFF")
        spinner_container.pack(pady=(10, 10))
        
        self.spinner_canvas = tk.Canvas(
            spinner_container,
            width=100,
            height=100,
            bg="#FFFFFF",
            highlightthickness=0
        )
        self.spinner_canvas.pack()
        
        self.spinner_angle = 0
        self.spinner_running = False
        
        self.status_message_label = tk.Label(
            self.progress_frame,
            text="Iniciando...",
            font=("Segoe UI", 12),
            fg=self.colors["button_bg"],
            bg="#FFFFFF",
        )
        self.status_message_label.pack(pady=(10, 15))

        logs_label = tk.Label(
            self.progress_frame,
            text="📋 Detalhes do processamento:",
            font=("Segoe UI", 11, "bold"),
            fg=self.colors["text_primary"],
            bg="#FFFFFF",
        )
        logs_label.pack(anchor="w", padx=24, pady=(15, 4))

        # Área de logs com scrollbar
        logs_container = tk.Frame(self.progress_frame, bg="#FFFFFF")
        logs_container.pack(fill="both", expand=True, padx=24, pady=(0, 20))

        self.logs_text = tk.Text(
            logs_container,
            height=12,
            wrap="word",
            font=("Consolas", 9),
            bg="#1E1E1E",
            fg="#D4D4D4",
            relief="flat",
            bd=0,
            padx=8,
            pady=8,
        )
        scrollbar = tk.Scrollbar(logs_container, command=self.logs_text.yview)
        self.logs_text.config(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        self.logs_text.pack(side="left", fill="both", expand=True)

    def _build_file_card(self, parent: tk.Frame, index: int, icon: str, title: str, subtitle: str) -> None:
        frame = tk.Frame(
            parent,
            bg="#F6F7F8",
            highlightbackground=self.colors["card_border"],
            highlightcolor=self.colors["card_border"],
            highlightthickness=2,
            bd=0,
            padx=14,
            pady=14,
        )
        frame.grid(row=0, column=index, padx=10, sticky="nsew")
        parent.columnconfigure(index, weight=1)

        icon_label = tk.Label(
            frame,
            text=icon,
            font=("Segoe UI Emoji", 44),
            fg="#1E2A34",
            bg="#F6F7F8",
        )
        icon_label.pack(pady=(6, 4))

        title_label = tk.Label(
            frame,
            text=title,
            font=("Segoe UI", 24, "bold"),
            fg=self.colors["text_primary"],
            bg="#F6F7F8",
        )
        title_label.pack(pady=(0, 2))

        subtitle_label = tk.Label(
            frame,
            text=subtitle,
            font=("Segoe UI", 15),
            fg=self.colors["text_secondary"],
            bg="#F6F7F8",
            wraplength=290,
            justify="center",
        )
        subtitle_label.pack(fill="x", pady=(0, 14))

        select_button = tk.Button(
            frame,
            text="Selecionar Arquivo",
            command=lambda idx=index: self.select_file(idx),
            font=("Segoe UI", 13, "bold"),
            bg=self.colors["button_bg"],
            fg=self.colors["button_text"],
            activebackground="#0E7CA2",
            activeforeground=self.colors["button_text"],
            relief="flat",
            bd=0,
            padx=18,
            pady=8,
            cursor="hand2",
        )
        select_button.pack(pady=(0, 10))

        file_label = tk.Label(
            frame,
            textvariable=self.file_name_vars[index],
            font=("Segoe UI", 10),
            fg="#0A5E7C",
            bg="#F6F7F8",
            wraplength=300,
            justify="center",
        )
        file_label.pack(fill="x")

    def _update_generate_button_state(self) -> None:
        ready = all(bool(var.get().strip()) for var in self.file_vars)
        if ready:
            self.generate_button.configure(state="normal", bg=self.colors["button_bg"])
        else:
            self.generate_button.configure(state="disabled", bg=self.colors["button_disabled"])

    def _set_status(self, message: str, *, status: str = "normal") -> None:
        self.status_var.set(message)
        if status == "error":
            color = self.colors["error"]
        elif status == "ok":
            color = self.colors["ok"]
        else:
            color = self.colors["text_primary"]
        self.status_label.configure(fg=color)

    def select_file(self, index: int) -> None:
        path = filedialog.askopenfilename(
            title=f"Selecione a {index + 1}ª planilha",
            filetypes=[("Planilhas Excel", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        if not path:
            return

        selected = Path(path).resolve()
        self.file_vars[index].set(selected.as_posix())
        self.file_name_vars[index].set(f"✓ {selected.name}")
        self._set_status(f"Arquivo {index + 1} selecionado: {selected.name}")
        self._update_generate_button_state()

    def generate(self) -> None:
        try:
            file_paths = []
            for index, var in enumerate(self.file_vars, start=1):
                raw_path = var.get().strip()
                if not raw_path:
                    raise ValueError(f"Selecione o Arquivo {index} antes de gerar.")

                path = Path(raw_path)
                validate_input_file(path)
                file_paths.append(path)

            default_name = f"Planilha_Consolidada_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            output = filedialog.asksaveasfilename(
                title="Salvar planilha consolidada",
                initialdir=str(Path.home() / "Downloads"),
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Planilhas Excel", "*.xlsx")],
            )
            if not output:
                self._set_status("Operação cancelada pelo usuário.")
                return

            # Preparar interface para processamento - CRIAR NOVA JANELA LIMPA
            self.generate_button.configure(state="disabled")
            
            # Esconder todos os widgets da janela principal, exceto o cabeçalho
            all_widgets = list(self.root.winfo_children())
            header_widget = all_widgets[0] if all_widgets else None
            
            for widget in all_widgets:
                if widget != header_widget:
                    widget.pack_forget()
            
            # Criar container principal para o progresso
            main_container = tk.Frame(self.root, bg="#08587A")
            main_container.pack(fill="both", expand=True)
            
            # Frame branco destacado para o progresso
            progress_card = tk.Frame(
                main_container,
                bg="#FFFFFF",
                highlightbackground="#1188B2",
                highlightthickness=3,
                relief="ridge"
            )
            progress_card.pack(fill="both", expand=True, padx=40, pady=40)
            
            # Título
            title_label = tk.Label(
                progress_card,
                text="⏳ Processando Planilhas",
                font=("Segoe UI", 20, "bold"),
                fg="#0C4560",
                bg="#FFFFFF"
            )
            title_label.pack(pady=(30, 20))
            
            # Canvas para spinner
            canvas = tk.Canvas(
                progress_card,
                width=120,
                height=120,
                bg="#FFFFFF",
                highlightthickness=0
            )
            canvas.pack(pady=20)
            
            # Label de status
            status_label = tk.Label(
                progress_card,
                text="Iniciando...",
                font=("Segoe UI", 13),
                fg="#1188B2",
                bg="#FFFFFF"
            )
            status_label.pack(pady=(10, 20))
            
            # Área de logs
            logs_frame = tk.Frame(progress_card, bg="#FFFFFF")
            logs_frame.pack(fill="both", expand=True, padx=30, pady=(10, 30))
            
            tk.Label(
                logs_frame,
                text="📋 Logs do Processamento:",
                font=("Segoe UI", 11, "bold"),
                fg="#0C4560",
                bg="#FFFFFF"
            ).pack(anchor="w", pady=(0, 5))
            
            logs_text = tk.Text(
                logs_frame,
                height=10,
                wrap="word",
                font=("Consolas", 9),
                bg="#1E1E1E",
                fg="#D4D4D4",
                relief="flat",
                bd=0,
                padx=10,
                pady=10
            )
            scrollbar = tk.Scrollbar(logs_frame, command=logs_text.yview)
            logs_text.config(yscrollcommand=scrollbar.set)
            scrollbar.pack(side="right", fill="y")
            logs_text.pack(fill="both", expand=True)
            
            # Forçar renderização
            self.root.update_idletasks()
            self.root.update()
            
            # Variáveis de controle do spinner
            spinner_data = {"angle": 0, "running": True}
            
            def animate_spinner():
                if not spinner_data["running"]:
                    return
                
                import math
                canvas.delete("all")
                
                center_x, center_y = 60, 60
                radius = 35
                dot_radius = 7
                num_dots = 12
                
                for i in range(num_dots):
                    angle = math.radians(spinner_data["angle"] + (i * (360 / num_dots)))
                    x = center_x + radius * math.cos(angle)
                    y = center_y + radius * math.sin(angle)
                    
                    opacity = 1.0 - (i / num_dots) * 0.85
                    r, g, b = 17, 136, 178  # #1188B2
                    
                    final_r = int(r * opacity + 255 * (1 - opacity))
                    final_g = int(g * opacity + 255 * (1 - opacity))
                    final_b = int(b * opacity + 255 * (1 - opacity))
                    
                    color = f'#{final_r:02x}{final_g:02x}{final_b:02x}'
                    
                    canvas.create_oval(
                        x - dot_radius, y - dot_radius,
                        x + dot_radius, y + dot_radius,
                        fill=color,
                        outline=""
                    )
                
                spinner_data["angle"] = (spinner_data["angle"] + 8) % 360
                
                if spinner_data["running"]:
                    self.root.after(50, animate_spinner)
            
            # Iniciar animação
            animate_spinner()
            
            # Atualizar novamente para garantir que o spinner apareça
            self.root.update()
            
            # Redirecionar stdout
            old_stdout = sys.stdout
            sys.stdout = TextRedirector(logs_text)
            
            print("=" * 60)
            print("PROCESSAMENTO INICIADO")
            print("=" * 60)
            print("")
            
            # Callback de progresso
            def update_progress_ui(percent: float, message: str):
                def update():
                    status_label.config(text=message)
                    print(f"[{int(percent)}%] {message}")
                self.root.after(0, update)
            
            # Executar em thread separada
            def process_in_background():
                try:
                    result = unify_spreadsheets(file_paths, Path(output).resolve(), update_progress_ui)
                    # Restaurar stdout
                    sys.stdout = old_stdout
                    
                    # Parar spinner e mostrar sucesso
                    def show_success():
                        spinner_data["running"] = False
                        canvas.delete("all")
                        
                        # Desenhar ícone de check
                        canvas.create_oval(30, 30, 90, 90, fill="#1D7F43", outline="")
                        canvas.create_text(60, 60, text="✓", font=("Arial", 40, "bold"), fill="white")
                        
                        status_label.config(text="✓ Processamento Concluído!", fg="#1D7F43")
                        print("\n" + "=" * 60)
                        print("PROCESSAMENTO CONCLUÍDO COM SUCESSO!")
                        print(f"Arquivo: {result}")
                        print("=" * 60)
                        
                        # Esperar 2 segundos e fechar
                        self.root.after(2000, lambda: [
                            messagebox.showinfo("Sucesso", f"Planilha consolidada salva em:\n{result}"),
                            self.root.quit()
                        ])
                    
                    self.root.after(0, show_success)
                    
                except Exception as error:
                    # Restaurar stdout
                    sys.stdout = old_stdout
                    
                    # Capturar o erro antes de definir a função interna
                    error_message = str(error)
                    
                    # Parar spinner e mostrar erro
                    def show_error():
                        spinner_data["running"] = False
                        canvas.delete("all")
                        
                        # Desenhar ícone de X
                        canvas.create_oval(30, 30, 90, 90, fill="#A11D2A", outline="")
                        canvas.create_text(60, 60, text="✗", font=("Arial", 40, "bold"), fill="white")
                        
                        status_label.config(text="✗ Erro no Processamento", fg="#A11D2A")
                        print("\n" + "=" * 60)
                        print("ERRO NO PROCESSAMENTO!")
                        print(f"Detalhes: {error_message}")
                        print("=" * 60)
                        
                        # Esperar 2 segundos e fechar
                        self.root.after(2000, lambda: [
                            messagebox.showerror("Erro", error_message),
                            self.root.quit()
                        ])
                    
                    self.root.after(0, show_error)
            
            # Iniciar thread após pequeno delay
            self.root.after(200, lambda: threading.Thread(target=process_in_background, daemon=True).start())
            
        except Exception as error:
            self._set_status(f"Erro: {error}", status="error")
            messagebox.showerror("Erro", str(error))
            self._update_generate_button_state()

    def _animate_spinner(self) -> None:
        """Anima o spinner de loading com círculos rotativos"""
        if not self.spinner_running:
            return
        
        import math
        
        # Limpar canvas
        self.spinner_canvas.delete("all")
        
        # Parâmetros do spinner
        center_x, center_y = 50, 50
        radius = 30
        dot_radius = 6
        num_dots = 12
        
        # Desenhar círculos com opacidade gradual
        for i in range(num_dots):
            angle = math.radians(self.spinner_angle + (i * (360 / num_dots)))
            x = center_x + radius * math.cos(angle)
            y = center_y + radius * math.sin(angle)
            
            # Calcular opacidade (o primeiro círculo é mais escuro)
            opacity = 1.0 - (i / num_dots) * 0.85
            
            # Converter cor base para RGB com opacidade
            base_color = self.colors["button_bg"].lstrip('#')
            r, g, b = int(base_color[0:2], 16), int(base_color[2:4], 16), int(base_color[4:6], 16)
            
            # Misturar com fundo branco baseado na opacidade
            bg_r, bg_g, bg_b = 255, 255, 255  # Branco
            
            final_r = int(r * opacity + bg_r * (1 - opacity))
            final_g = int(g * opacity + bg_g * (1 - opacity))
            final_b = int(b * opacity + bg_b * (1 - opacity))
            
            color = f'#{final_r:02x}{final_g:02x}{final_b:02x}'
            
            self.spinner_canvas.create_oval(
                x - dot_radius, y - dot_radius,
                x + dot_radius, y + dot_radius,
                fill=color,
                outline=""
            )
        
        # Incrementar ângulo para próximo frame
        self.spinner_angle = (self.spinner_angle + 8) % 360
        
        # Agendar próximo frame (20 FPS para suavidade)
        if self.spinner_running:
            self.root.after(50, self._animate_spinner)
    
    def _on_success(self, result: Path) -> None:
        self.spinner_running = False
        self.status_message_label.config(text="✓ Concluído com sucesso!", fg=self.colors["ok"])
        self._set_status(f"✓ Concluído com sucesso: {result}", status="ok")
        self._update_generate_button_state()
        
        # Adicionar mensagem final aos logs
        print("\n" + "=" * 60)
        print("PROCESSAMENTO CONCLUÍDO COM SUCESSO!")
        print(f"Arquivo salvo em: {result}")
        print("=" * 60)
        
        messagebox.showinfo("Sucesso", f"Planilha consolidada salva em:\n{result}")

    def _on_error(self, error: Exception) -> None:
        self.spinner_running = False
        self.status_message_label.config(text="✗ Erro no processamento", fg=self.colors["error"])
        self._set_status(f"✗ Erro: {error}", status="error")
        self._update_generate_button_state()
        
        # Adicionar mensagem de erro aos logs
        print("\n" + "=" * 60)
        print("ERRO NO PROCESSAMENTO!")
        print(f"Detalhes: {error}")
        print("=" * 60)
        
        messagebox.showerror("Erro", str(error))


def run_gui() -> int:
    root = tk.Tk()
    UnificadorApp(root)
    root.mainloop()
    return 0


def main() -> int:
    args = parse_args()

    if not args.cli and not args.files:
        return run_gui()

    try:
        input_files = prompt_for_files_if_needed(args.files)
        for file_path in input_files:
            validate_input_file(file_path)

        output_path = build_output_path(args.output)
        result = unify_spreadsheets(input_files, output_path)

        print("\n✓ Unificação concluída com sucesso")
        print(f"Arquivo final: {result}")
        return 0
    except Exception as error:
        print(f"\n✗ Erro: {error}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
